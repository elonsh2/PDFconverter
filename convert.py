from __future__ import print_function
from mailmerge import MailMerge
import openpyxl
from docx2pdf import convert
import os
from PyPDF2 import PdfFileMerger


class Convert:
	def __init__(self, convert_path, excel_file_path, word_file, final_path,
				 appendix_file):
		"""
		:param convert_path: the directory where the converted PDFs will be
		:param excel_file_path: the path of the Excel file of the case numbers
		:param word_file: the path of the Word file to fill and duplicate
		:param final_path: the directory where the merged PDFs will be
		:param appendix_file: PDF file to add to each converted PDF
		"""
		self.appendix_file = appendix_file
		self.final_path = final_path
		self.word_file = word_file
		self.excel_file_path = excel_file_path
		self.convert_path = convert_path

	def create_word(self):
		"""
		this function create word files from the data in the Excel file
		:return: none
		"""
		work_book = openpyxl.load_workbook(self.excel_file_path)
		sh = work_book.active
		template_1 = self.word_file
		document_1 = MailMerge(template_1)
		print("Fields included in {}: {}".format(template_1,
												 document_1.get_merge_fields()))
		# iterate through excel and display data
		for i in range(1, sh.max_row + 1):  # sh.max_row + 1
			for j in range(1, sh.max_column + 1):
				cell_obj = sh.cell(row=i, column=j)
				document_1 = MailMerge(template_1)
				document_1.merge(
					case_number=f"{cell_obj.value[0:-1]}")
				document_1.write(
					f"{self.convert_path}{cell_obj.value[0:-1]}.docx")

	def convert_to_pdf(self):
		"""
		this function converts the Word files to PDFs
		:return: none
		"""
		convert(self.convert_path)
		test = os.listdir(self.convert_path)
		for item in test:
			if item.endswith(".docx"):
				os.remove(os.path.join(self.convert_path, item))

	def merge_pdf(self):
		pdfs = [a for a in os.listdir(self.convert_path)]
		for i, pdf in enumerate(pdfs):
			merger = PdfFileMerger()
			merger.append(open(f"{self.convert_path}{pdf}", 'rb'))
			merger.append(open(self.appendix_file, 'rb'))
			# sets the name of each combined PDF
			with open(
					f"{self.final_path}בקשת עיון בתיק {pdf[:-4]} שרון שקרג'י.pdf",
					"wb") as fout:
				merger.write(fout)
