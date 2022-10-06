from __future__ import print_function
from mailmerge import MailMerge
import openpyxl
from docx2pdf import convert
import os
from PyPDF2 import PdfFileMerger
from convert import *

# convert_path: the directory where the converted PDFs will be
convert_path = "C:\\Users\\Elon\\Desktop\\Exceltoword\\work\\convert\\"
# final_path: the directory where the merged PDFs will be
final_path = "C:\\Users\\Elon\\Desktop\\Exceltoword\\work\\combined\\"
# excel_file_path: the path of the Excel file of the case numbers
excel_file_path = "C:\\Users\\Elon\\Desktop\\Exceltoword\\test.xlsx"
# word_file = the path of the Word file to fill and duplicate
word_file = "C:\\Users\\Elon\\Desktop\\Exceltoword\\to_fill.docx"
# appendix_file: PDF file to add to each converted PDF
appendix_file = "C:\\Users\\Elon\\Desktop\\Exceltoword\\zzz.pdf"

# def create_word():
# 	wrkbk = openpyxl.load_workbook(excel_file_path)
# 	sh = wrkbk.active
# 	template_1 = word_file
# 	document_1 = MailMerge(template_1)
# 	print("Fields included in {}: {}".format(template_1,
# 											 document_1.get_merge_fields()))
# 	# iterate through excel and display data
# 	for i in range(1, sh.max_row + 1):  # sh.max_row + 1
# 		for j in range(1, sh.max_column + 1):
# 			cell_obj = sh.cell(row=i, column=j)
# 			document_1 = MailMerge(template_1)
# 			document_1.merge(
# 				case_number=f"{cell_obj.value[0:-1]}")
# 			document_1.write(f"{convert_path}{cell_obj.value[0:-1]}.docx")
#
#
# def convert_to_pdf():
# 	convert(convert_path)
# 	test = os.listdir(convert_path)
# 	for item in test:
# 		if item.endswith(".docx"):
# 			os.remove(os.path.join(convert_path, item))
#
#
# def merge_pdf():
# 	pdfs = [a for a in os.listdir(convert_path)]
# 	for i, pdf in enumerate(pdfs):
# 		merger = PdfFileMerger()
# 		merger.append(open(f"{convert_path}{pdf}", 'rb'))
# 		merger.append(open("zzz.pdf", 'rb'))
# 		with open(f"{final_path}בקשת עיון בתיק {pdf[:-4]} שרון שקרג'י.pdf",
# 				  "wb") as fout:
# 			merger.write(fout)


if __name__ == "__main__":
	convert = Convert(convert_path, excel_file_path, word_file,
					  final_path, appendix_file)
	convert.create_word()
	convert.convert_to_pdf()
	convert.merge_pdf()

